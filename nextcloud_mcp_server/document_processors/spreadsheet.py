"""Spreadsheets, read cell-by-cell rather than through a PDF rendition.

Deliberately NOT the rendition route the word-processor formats take. Measured
on a real workbook, rendering to PDF paginated it into 25 mixed-orientation
pages and recalled only **63.8%** of the tokens a direct cell read recovers --
LibreOffice honours the print layout, so columns past the print width are cut
and the question numbering (``1.2``, ``2.a``, ``3.a``..``3.d``) disappears
entirely. A direct read recovered 2013 cells against the rendition's 567.

A spreadsheet also has no page geometry to highlight, so giving up the
rendition costs nothing a viewer could have used: chunks carry a sheet name and
cell range instead of a bounding box.
"""

import io
import logging
from collections.abc import Awaitable, Callable
from typing import Any, Optional

from anyio.to_thread import run_sync

from . import _libreoffice
from .base import DocumentProcessor, ProcessingResult, ProcessorError

logger = logging.getLogger(__name__)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLS_MIME = "application/vnd.ms-excel"
SPREADSHEET_MIME_TYPES = {XLSX_MIME, XLS_MIME}

# Per-sheet spans into the joined text: ``{"sheet", "cell_range",
# "start_offset", "end_offset"}``. The spreadsheet counterpart to a PDF's
# ``page_boundaries`` -- it lets a chunk be attributed to the sheet and cell
# range it came from, which is the anchor a viewer can use where there is no
# bbox to draw.
SHEET_BOUNDARIES_KEY = "sheet_boundaries"


class SpreadsheetProcessor(DocumentProcessor):
    """Extract ``.xlsx``/``.xls`` as one markdown table per sheet."""

    def __init__(self, timeout: float = 120.0):
        self._timeout = timeout

    @property
    def name(self) -> str:
        return "spreadsheet"

    @property
    def tier(self) -> str:
        return "fast"

    @property
    def supported_mime_types(self) -> set[str]:
        # Legacy .xls needs LibreOffice to reach a readable container. Where it
        # is absent, claim only .xlsx: the registry then reports "no processor
        # for application/vnd.ms-excel" once, instead of this processor
        # accepting every .xls and failing on each one.
        if _libreoffice.LIBREOFFICE_AVAILABLE:
            return SPREADSHEET_MIME_TYPES
        return {XLSX_MIME}

    async def process(
        self,
        content: bytes,
        content_type: str,
        filename: Optional[str] = None,
        options: Optional[dict[str, Any]] = None,
        progress_callback: Optional[
            Callable[[float, Optional[float], Optional[str]], Awaitable[None]]
        ] = None,
    ) -> ProcessingResult:
        base_type = content_type.split(";")[0].strip().lower()
        converted_from = None

        if base_type == XLS_MIME:
            # openpyxl cannot read the legacy OLE2 format. Convert the container
            # to xlsx rather than to PDF: this is a format change, not a
            # re-layout, so it keeps every cell instead of losing a third of
            # them to pagination.
            if progress_callback:
                await progress_callback(0.0, None, "Converting legacy spreadsheet...")
            try:
                content = await _libreoffice.convert(
                    content,
                    filename or "workbook.xls",
                    "xlsx",
                    timeout_seconds=self._timeout,
                )
            except _libreoffice.LibreOfficeError as exc:
                raise ProcessorError(
                    f"Legacy spreadsheet conversion failed: {exc}"
                ) from exc
            converted_from = XLS_MIME

        try:
            text, boundaries, sheet_names = await run_sync(_extract_workbook, content)
        except ProcessorError:
            raise
        except Exception as exc:
            raise ProcessorError(f"Spreadsheet parse failed: {exc}") from exc

        metadata: dict[str, Any] = {
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            SHEET_BOUNDARIES_KEY: boundaries,
            "text_length": len(text),
            "parse_mode": "markdown",
        }
        if converted_from:
            metadata["converted_from_mime"] = converted_from

        return ProcessingResult(
            text=text,
            metadata=metadata,
            processor=self.name,
            success=True,
        )

    async def health_check(self) -> bool:
        try:
            import openpyxl  # noqa: F401, PLC0415
        except ImportError:
            return False
        return True


def _escape(value: Any) -> str:
    """One cell as markdown-table-safe text.

    A literal ``|`` in a cell would end the column early and shift every value
    after it into the wrong heading, so it is escaped; newlines inside a cell
    would end the row entirely, so they become spaces.
    """
    if value is None:
        return ""
    text = str(value).replace("|", "\\|")
    return " ".join(text.split())


def _extract_workbook(content: bytes) -> tuple[str, list[dict[str, Any]], list[str]]:
    """Render every sheet as a markdown table. Runs in a worker thread."""
    import openpyxl  # noqa: PLC0415 -- keep the import off the hot path

    # data_only: read the cached result of a formula rather than "=SUM(A1:A9)",
    # which is what a reader searches for. read_only streams rows instead of
    # building the whole object graph, so a large workbook stays bounded.
    workbook = openpyxl.load_workbook(
        io.BytesIO(content), data_only=True, read_only=True
    )
    try:
        parts: list[str] = []
        boundaries: list[dict[str, Any]] = []
        offset = 0
        for sheet in workbook.worksheets:
            rows = _sheet_rows(sheet)
            if not rows:
                continue
            body = f"## {sheet.title}\n\n" + "\n".join(rows) + "\n"
            parts.append(body)
            boundaries.append(
                {
                    "sheet": sheet.title,
                    "cell_range": sheet.calculate_dimension(),
                    "start_offset": offset,
                    "end_offset": offset + len(body),
                }
            )
            offset += len(body)
        return "".join(parts), boundaries, workbook.sheetnames
    finally:
        workbook.close()


def _sheet_rows(sheet: Any) -> list[str]:
    """Markdown rows for one sheet, or [] when it holds nothing.

    The first non-empty row becomes the header. That is a guess -- a real
    spreadsheet may open with a title banner -- but the alternative is an empty
    header row, and the cells are all present either way, which is what a
    retrieval index needs.
    """
    rendered: list[list[str]] = []
    width = 0
    for row in sheet.iter_rows(values_only=True):
        cells = [_escape(v) for v in row]
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue
        width = max(width, len(cells))
        rendered.append(cells)

    if not rendered:
        return []

    lines = []
    for i, cells in enumerate(rendered):
        padded = cells + [""] * (width - len(cells))
        lines.append("| " + " | ".join(padded) + " |")
        if i == 0:
            lines.append("| " + " | ".join(["---"] * width) + " |")
    return lines
